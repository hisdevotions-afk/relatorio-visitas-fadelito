import os
import sys
import json
import time
from datetime import datetime, date, timedelta
from pathlib import Path

import requests
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import PatternFill

FILL_AMARELO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

load_dotenv()

API_BASE_URL  = os.getenv("API_BASE_URL", "").rstrip("/")
API_TOKEN     = os.getenv("API_TOKEN", "")
API_TIMEOUT   = int(os.getenv("API_TIMEOUT", "30"))
DATA_INICIO   = os.getenv("DATA_INICIO", "")   # YYYY-MM-DD; usa hoje se vazio
DATA_FIM      = os.getenv("DATA_FIM", "")      # YYYY-MM-DD; usa DATA_INICIO se vazio
OUTPUT_DIR    = os.getenv("OUTPUT_DIR", ".")

STATUS_LABELS   = {"1": "Agendado", "2": "Confirmado", "3": "Realizado", "7": "Cancelado", "9": "Faltou"}
STATUS_VALIDOS  = set(STATUS_LABELS)
STATUS_PENDENTE = {"Agendado", "Confirmado"}  # linhas que receberão destaque amarelo

COLUNAS_EXCEL = ["ID", "Data/Hora", "Paciente", "Atendente", "Serviço", "Status", "Telefone"]


def _validate_env() -> None:
    missing = [k for k in ("API_BASE_URL", "API_TOKEN") if not os.getenv(k)]
    if missing:
        sys.exit(f"[ERRO] Variáveis de ambiente obrigatórias não definidas: {', '.join(missing)}")


def _last_business_day() -> date:
    today = date.today()
    # Monday (weekday=0) → back 3 days to Friday; otherwise → yesterday
    delta = 3 if today.weekday() == 0 else 1
    return today - timedelta(days=delta)


def _resolve_dates() -> tuple[str, str]:
    def parse(value: str, label: str) -> str:
        try:
            datetime.strptime(value, "%Y-%m-%d")
            return value
        except ValueError:
            sys.exit(f"[ERRO] {label} inválida: '{value}'. Use o formato YYYY-MM-DD.")

    inicio = parse(DATA_INICIO, "DATA_INICIO") if DATA_INICIO else _last_business_day().isoformat()
    fim    = parse(DATA_FIM, "DATA_FIM")       if DATA_FIM    else inicio
    return inicio, fim


def _fetch_agendamentos(inicio: str, fim: str) -> list[dict]:
    endpoint = f"{API_BASE_URL}/agendamentos"
    headers  = {"Authorization": f"Bearer {API_TOKEN}", "Accept": "application/json"}
    params   = {"Inicio": inicio, "Fim": fim}

    try:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=API_TIMEOUT)
        resp.raise_for_status()
    except requests.exceptions.ConnectionError:
        sys.exit(f"[ERRO] Não foi possível conectar à API: {endpoint}")
    except requests.exceptions.Timeout:
        sys.exit(f"[ERRO] Timeout após {API_TIMEOUT}s ao chamar a API.")
    except requests.exceptions.HTTPError as exc:
        sys.exit(f"[ERRO] API retornou HTTP {resp.status_code}: {exc}")

    try:
        payload = resp.json()
    except json.JSONDecodeError:
        sys.exit("[ERRO] Resposta da API não é JSON válido.")

    if isinstance(payload, dict):
        if "data" not in payload:
            sys.exit("[ERRO] Chave 'data' não encontrada na resposta da API.")
        payload = payload["data"]

    if not isinstance(payload, list):
        sys.exit(
            f"[ERRO] Esperava lista em 'data', mas recebeu {type(payload).__name__}."
        )

    return payload


def _build_dataframe(agendamentos: list[dict]) -> pd.DataFrame:
    if not agendamentos:
        return pd.DataFrame(columns=COLUNAS_EXCEL)

    rows = []
    for a in agendamentos:
        if a.get("id") == 0:
            continue
        if a.get("status_agendamento") not in STATUS_VALIDOS:
            continue
        if str(a.get("servico") or "").strip().lower() != "visita":
            continue
        rows.append({
            "ID":        a.get("id"),
            "Data/Hora": _fmt_datetime(a.get("start")),
            "Paciente":  a.get("cliente"),
            "Atendente": a.get("atendente"),
            "Serviço":   a.get("servico"),
            "Status":    STATUS_LABELS[a.get("status_agendamento")],
            "Telefone":  None,
        })

    return pd.DataFrame(rows, columns=COLUNAS_EXCEL)


def _fmt_datetime(value: str | None) -> str | None:
    if not value:
        return None
    try:
        return datetime.strptime(value[:19], "%Y-%m-%dT%H:%M:%S").strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return value


def _enrich_telefones(df: pd.DataFrame) -> None:
    headers = {"Authorization": f"Bearer {API_TOKEN}", "Accept": "application/json"}
    total   = len(df)

    for i, (idx, row) in enumerate(df.iterrows(), start=1):
        agendamento_id = row["ID"]
        print(f"[INFO] Buscando telefone {i}/{total} (ID {agendamento_id})...", end="\r")

        try:
            resp = requests.get(
                f"{API_BASE_URL}/agendamento-dados/{agendamento_id}",
                headers=headers,
                timeout=API_TIMEOUT,
            )
            resp.raise_for_status()
            dado = resp.json()

            # Navega até o objeto de dados se a resposta vier envelopada em 'data'
            if isinstance(dado, dict) and "data" in dado:
                dado = dado["data"]

            telefone = dado.get("paciente_telefone") if isinstance(dado, dict) else None
            df.at[idx, "Telefone"] = telefone

        except Exception as exc:
            print(f"\n[AVISO] Não foi possível buscar telefone do ID {agendamento_id}: {exc}")

        if i < total:
            time.sleep(0.3)

    print()  # quebra a linha do \r


def _save_excel(df: pd.DataFrame, inicio: str, fim: str) -> Path:
    output_path = Path(OUTPUT_DIR)
    output_path.mkdir(parents=True, exist_ok=True)

    suffix = inicio.replace("-", "")
    if fim != inicio:
        suffix += f"_{fim.replace('-', '')}"
    filepath = output_path / f"agendamentos_{suffix}.xlsx"

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Agendamentos")

        ws = writer.sheets["Agendamentos"]
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

        status_col_idx = COLUNAS_EXCEL.index("Status") + 1  # openpyxl é 1-based
        num_cols = len(COLUNAS_EXCEL)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=num_cols):
            status_cell = row[status_col_idx - 1]
            if status_cell.value in STATUS_PENDENTE:
                for cell in row:
                    cell.fill = FILL_AMARELO

    return filepath


def main() -> None:
    _validate_env()

    inicio, fim = _resolve_dates()
    periodo = inicio if inicio == fim else f"{inicio} a {fim}"
    print(f"[INFO] Consultando agendamentos: {periodo}")

    agendamentos = _fetch_agendamentos(inicio, fim)
    total_bruto  = len(agendamentos)

    df    = _build_dataframe(agendamentos)
    total = len(df)

    if total == 0:
        print("[INFO] Nenhum agendamento válido encontrado para o período consultado. "
              "O arquivo Excel será gerado apenas com o cabeçalho.")
    else:
        print(f"[INFO] Buscando telefones ({total} agendamentos)...")
        _enrich_telefones(df)

    filepath = _save_excel(df, inicio, fim)

    print("-" * 50)
    print(f"Período consultado   : {periodo}")
    print(f"Registros da API     : {total_bruto}")
    print(f"Agendamentos válidos : {total}  (apenas 'Visita'; excluídos bloqueios e status fora de 1/2/3/7/9)")
    print(f"Arquivo gerado       : {filepath.resolve()}")
    print("-" * 50)


if __name__ == "__main__":
    main()
